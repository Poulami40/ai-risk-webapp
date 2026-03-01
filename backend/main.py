from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
import google.generativeai as genai
import pdfplumber
import docx
import json
from openpyxl import load_workbook
import uuid
import os

app = FastAPI()

# ---------- GEMINI SETUP ----------
genai.configure(api_key="AIzaSyDyznawf3wjkq1ibWST_bQJm3INADMYbiA")

model = genai.GenerativeModel("gemini-1.5-flash")

# ---------- TEXT EXTRACTION ----------
def extract_text(file_path):
    if file_path.endswith(".pdf"):
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
        return text

    if file_path.endswith(".docx"):
        doc = docx.Document(file_path)
        return "\n".join([p.text for p in doc.paragraphs])

    return ""


# ---------- GEMINI ANALYSIS ----------
def analyze_document(text):

    prompt = f"""
    Extract information from the document and return ONLY JSON.

    Required fields:
    company_name
    industry
    data_sensitivity
    risk_type
    controls_present

    If missing return null.

    Document:
    {text}
    """

    response = model.generate_content(prompt)

    return json.loads(response.text)


# ---------- EXCEL FILL ----------
def fill_excel(data):

    template_path = "demo-template.xlsx"
    output_path = f"output_{uuid.uuid4()}.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    ws["B2"] = data.get("company_name")
    ws["B3"] = data.get("industry")
    ws["B4"] = data.get("data_sensitivity")
    ws["B5"] = data.get("risk_type")
    ws["B6"] = data.get("controls_present")

    wb.save(output_path)

    return output_path


# ---------- API ----------
@app.post("/analyze")
async def analyze(file: UploadFile = File(...)):

    temp_file = f"temp_{file.filename}"

    with open(temp_file, "wb") as f:
        f.write(await file.read())

    text = extract_text(temp_file)

    extracted_data = analyze_document(text)

    excel_file = fill_excel(extracted_data)

    return FileResponse(
        excel_file,
        filename="filled_assessment.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )